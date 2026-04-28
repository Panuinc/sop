"""สร้าง HTML ฉบับเดียวรวมทุกอย่าง — กดที่ box แล้วเปิด popup รายละเอียดเต็ม
   อ่านข้อมูลจาก Excel Master Workflow (Detailed) เป็น Single Source of Truth"""

from openpyxl import load_workbook
import json

# ============ Read Excel data ============
wb = load_workbook("CHH_SOP update.xlsx")
ws = wb["Master Workflow (Detailed)"]

all_steps = []
current_section = None
step_num = 0

for row in ws.iter_rows(min_row=5, values_only=True):
    if all(c is None for c in row):
        continue
    cell0 = row[0]
    if cell0 is None:
        continue
    rest_empty = all(v is None for v in row[1:])
    if rest_empty:
        current_section = str(cell0).strip()
        continue
    if row[1]:
        step_num += 1
        all_steps.append({
            "num": step_num,
            "sop": str(cell0).strip(),
            "section": current_section or "",
            "title": str(row[1]),
            "tool": str(row[2]) if row[2] else "—",
            "input": str(row[3]) if row[3] else "—",
            "data": str(row[4]) if row[4] else "—",
            "process": str(row[5]) if row[5] else "—",
            "output": str(row[6]) if row[6] else "—",
            "role": str(row[7]) if row[7] else "—",
            "approver": str(row[8]) if row[8] else "—",
            "note": str(row[9]) if row[9] else "—",
        })

print(f"Read {len(all_steps)} steps from Excel")

# ============ Classify kind ============
CONTROL_POINTS = {
    "A2","A7","B3","B9","C4","C10","I2","D3","D11","E1","F10","H5",
    "J3","J4","K1","K4","L5","N2","N5","P4","V1","V2"
}
GATES = {"B12","F3"}
ARTIFACTS = {"B7"}

def kind_of(sop):
    if sop in GATES: return "gate"
    if sop in CONTROL_POINTS: return "ctrl"
    if sop in ARTIFACTS: return "artifact"
    return "proc"

for s in all_steps:
    s["kind"] = kind_of(s["sop"])

# Split into operations (Phase 1-6, sop in M/A/B/C/I/D/E/F/G/H) vs support (J,K,L,N,O,P,Q,R,S,T,U,V,W)
SUPPORT_PREFIXES = {"J","K","L","N","O","P","Q","R","S","T","U","V","W"}
ops_steps = [s for s in all_steps if s["sop"][0] not in SUPPORT_PREFIXES]
sup_steps = [s for s in all_steps if s["sop"][0] in SUPPORT_PREFIXES]

print(f"Operations: {len(ops_steps)} | Support: {len(sup_steps)}")

# ============ Helpers ============
def esc(s):
    return str(s).replace("&","&amp;").replace("<","&lt;").replace(">","&gt;")

def text(x, y, s, size=11, weight="normal", anchor="middle", color="#1a2332"):
    return f'<text x="{x}" y="{y}" text-anchor="{anchor}" font-size="{size}" font-weight="{weight}" fill="{color}" font-family="Sarabun, Arial, sans-serif">{esc(s)}</text>'

def wrap_thai(s, max_len=33):
    if len(s) <= max_len:
        return [s]
    breaks = [' ','/','+','·','—']
    for i in range(min(max_len, len(s)-1), max(max_len-15, 0), -1):
        if s[i] in breaks:
            return [s[:i].strip(), s[i+1:].strip() if len(s[i+1:]) <= max_len else s[i+1:i+1+max_len].strip()+"…"]
    return [s[:max_len], s[max_len:max_len*2]]

def render_box(x, y, w, h, step, marker_text="", header_color="#5D4037", title_size=12.5):
    """Clickable step box - wraps content in <g data-step="N">"""
    kind = step["kind"]
    if kind == "ctrl":
        fill, stroke, sw = "#FFF9C4","#F57F17",3
        marker = marker_text or "⚑ จุดตรวจสอบ"
        head_color = "#E65100"
    elif kind == "gate":
        fill, stroke, sw = "#FFEBEE","#C62828",3
        marker = marker_text or "⛔ ด่านห้ามข้าม"
        head_color = "#B71C1C"
    elif kind == "artifact":
        fill, stroke, sw = "#E1BEE7","#6A1B9A",1.5
        marker = marker_text or "📋 เอกสารหลัก"
        head_color = "#4A148C"
    else:
        fill, stroke, sw = "#FFF3CD","#A47D14",1.5
        marker = ""
        head_color = header_color

    out = [f'<g class="step-box" data-step="{step["num"]}" style="cursor:pointer">']
    rx = 22 if kind == "artifact" else 4
    out.append(f'<rect x="{x}" y="{y}" width="{w}" height="{h}" fill="{fill}" stroke="{stroke}" stroke-width="{sw}" rx="{rx}"/>')
    out.append(f'<rect x="{x}" y="{y}" width="{w}" height="22" fill="rgba(0,0,0,0.07)" rx="{rx}"/>')
    out.append(text(x+8, y+15, f"ขั้นที่ {step['num']} · {step['sop']}", 10.5, "bold", "start", head_color))
    if marker:
        out.append(text(x+w-8, y+15, marker, 9.5, "bold", "end", head_color))
    # title
    lines = wrap_thai(step["title"], 33)
    cx = x + w/2
    if len(lines) == 1:
        out.append(text(cx, y+44, lines[0], title_size, "bold", "middle", "#1a2332"))
    else:
        out.append(text(cx, y+38, lines[0], title_size, "bold", "middle", "#1a2332"))
        out.append(text(cx, y+54, lines[1], title_size, "bold", "middle", "#1a2332"))
    # role + approver
    role_text = f"👤 {step['role']}"
    if step["approver"] and step["approver"] != "—":
        role_text += f"   ✍ {step['approver']}"
    if len(role_text) > 50:
        role_text = role_text[:48]+"…"
    out.append(text(x+10, y+74, role_text, 10, "normal", "start", "#37474F"))
    # tool
    tool_text = f"🛠 {step['tool']}"
    if len(tool_text) > 52:
        tool_text = tool_text[:50]+"…"
    out.append(text(x+10, y+90, tool_text, 10, "normal", "start", "#546E7A"))
    # note
    if step["note"] and step["note"] != "—":
        note_short = step["note"] if len(step["note"]) <= 50 else step["note"][:48]+"…"
        out.append(text(x+10, y+105, f"💡 {note_short}", 9.5, "normal", "start", "#78909C"))
    out.append('</g>')
    return out

def arrow_v(x, y1, y2, color="#2c3e50"):
    return f'<path d="M {x},{y1} L {x},{y2}" stroke="{color}" stroke-width="1.8" fill="none" marker-end="url(#arr)"/>'

# ========================================
# SVG #1: Operations Swimlane (Phase 1-6)
# ========================================
W1 = 2050
PAD_L = 25
COL_W = 333
HEADER_H = 95
BOX_W = 305
BOX_H = 112
BOX_GAP = 12
START_Y = 135

phase_groups = {1:[],2:[],3:[],4:[],5:[],6:[]}
phase_names = [
    "Phase 1: เริ่มต้น + ทำราคา",
    "Phase 2: เขียนแบบ + เปิดออเดอร์",
    "Phase 3: วางแผน + จัดซื้อ + ตรวจรับวัสดุ",
    "Phase 4: ผลิต",
    "Phase 5: ตรวจคุณภาพ + แพ็ค + ส่ง",
    "Phase 6: เคลม + เก็บเงิน",
]
for s in ops_steps:
    p = s["sop"][0]
    if p == "M": phase_groups[1].append(s)
    elif p == "A": phase_groups[1].append(s)
    elif p == "B": phase_groups[2].append(s)
    elif p == "C": phase_groups[3].append(s)
    elif p == "I": phase_groups[3].append(s)
    elif p == "D": phase_groups[4].append(s)
    elif p == "E": phase_groups[5].append(s)
    elif p == "F": phase_groups[5].append(s)
    elif p == "G": phase_groups[6].append(s)
    elif p == "H": phase_groups[6].append(s)

# Re-order phase 3 to put C10 last (after I)
phase_groups[3] = [s for s in phase_groups[3] if s["sop"] != "C10"] + [s for s in phase_groups[3] if s["sop"] == "C10"]

max_steps_per_phase = max(len(g) for g in phase_groups.values())
H1 = START_Y + max_steps_per_phase * (BOX_H + BOX_GAP) + 80

elements = []
elements.append(f'<rect x="0" y="0" width="{W1}" height="{H1}" fill="#F5F8FC"/>')
elements.append(f'<rect x="0" y="0" width="{W1}" height="{HEADER_H}" fill="#1F4E78"/>')
elements.append(text(W1/2, 28, "งานประจำวัน 90 ขั้น (Phase 1-6)", 22, "bold", "middle", "#FFFFFF"))
elements.append(text(W1/2, 52, "เดินตามลำดับขั้นที่ 1 → 90 ตามลูกศร · กดที่กล่องใดๆ เพื่อดูรายละเอียดเต็ม", 12, "normal", "middle", "#D6E4F0"))
elements.append(text(W1/2, 72, "🟨 ปกติ · 🟧 จุดตรวจสอบ ⚑ · 🟥 ด่านห้ามข้าม ⛔ · 🟪 เอกสารหลัก 📋", 11, "normal", "middle", "#D6E4F0"))

for phase_idx in range(6):
    col_x = PAD_L + phase_idx * COL_W
    box_x = col_x + (COL_W - BOX_W) / 2
    bg = "#FAFCFE" if phase_idx % 2 == 0 else "#F0F5FA"
    elements.append(f'<rect x="{col_x}" y="{HEADER_H}" width="{COL_W}" height="{H1-HEADER_H}" fill="{bg}"/>')
    elements.append(f'<rect x="{col_x}" y="{HEADER_H}" width="{COL_W}" height="40" fill="#2E75B6" stroke="#1F4E78"/>')
    elements.append(text(col_x + COL_W/2, HEADER_H+26, phase_names[phase_idx], 14, "bold", "middle", "#FFFFFF"))
    if phase_idx < 5:
        elements.append(f'<line x1="{col_x+COL_W}" y1="0" x2="{col_x+COL_W}" y2="{H1}" stroke="#90A8C5" stroke-width="1.5"/>')

    box_positions = []
    for row_idx, step in enumerate(phase_groups[phase_idx+1]):
        y = START_Y + row_idx * (BOX_H + BOX_GAP)
        elements.extend(render_box(box_x, y, BOX_W, BOX_H, step))
        box_positions.append((box_x, y))
    for i in range(len(box_positions)-1):
        bx, by = box_positions[i]
        _, by_next = box_positions[i+1]
        elements.append(arrow_v(bx + BOX_W/2, by + BOX_H, by_next))
    # cross-phase arrow
    if phase_idx < 5 and box_positions:
        last_bx, last_by = box_positions[-1]
        next_col_x = PAD_L + (phase_idx+1) * COL_W
        next_box_x = next_col_x + (COL_W - BOX_W) / 2
        x1 = last_bx + BOX_W
        y1 = last_by + BOX_H/2
        x2 = next_box_x + BOX_W/2
        midy = HEADER_H + 60
        elements.append(f'<path d="M {x1},{y1} L {x1+8},{y1} L {x1+8},{midy} L {x2},{midy} L {x2},{START_Y}" stroke="#1F4E78" stroke-width="2.5" fill="none" marker-end="url(#arrB)"/>')

# CAPA loop
g8 = next((s for s in ops_steps if s["sop"]=="G8"), None)
if g8:
    g8_phase_idx = 5
    g8_row_idx = phase_groups[6].index(g8)
    g8_y = START_Y + g8_row_idx * (BOX_H + BOX_GAP) + BOX_H/2
    g8_x = PAD_L + g8_phase_idx * COL_W + (COL_W - BOX_W)/2
    d11 = next((s for s in ops_steps if s["sop"]=="D11"), None)
    if d11:
        d11_row_idx = phase_groups[4].index(d11)
        d11_y = START_Y + d11_row_idx * (BOX_H + BOX_GAP) + BOX_H/2
        d11_x = PAD_L + 3 * COL_W + (COL_W - BOX_W)/2 + BOX_W
        loop_y = H1 - 30
        elements.append(f'<path d="M {g8_x},{g8_y+30} L {g8_x-15},{g8_y+30} L {g8_x-15},{loop_y} L {d11_x+30},{loop_y} L {d11_x+30},{d11_y} L {d11_x},{d11_y}" stroke="#7c8a9a" stroke-width="2" fill="none" stroke-dasharray="8 5" marker-end="url(#arrG)"/>')
        elements.append(text((g8_x + d11_x)/2, loop_y - 8, "↩ CAPA Feedback Loop", 12, "bold", "middle", "#546E7A"))

svg_ops = f'<svg viewBox="0 0 {W1} {H1}" xmlns="http://www.w3.org/2000/svg" preserveAspectRatio="xMidYMid meet">\n' \
          '<defs>\n' \
          '<marker id="arr" markerWidth="9" markerHeight="9" refX="8" refY="3" orient="auto" markerUnits="strokeWidth"><path d="M0,0 L0,6 L9,3 Z" fill="#2c3e50"/></marker>\n' \
          '<marker id="arrB" markerWidth="9" markerHeight="9" refX="8" refY="3" orient="auto" markerUnits="strokeWidth"><path d="M0,0 L0,6 L9,3 Z" fill="#1F4E78"/></marker>\n' \
          '<marker id="arrG" markerWidth="9" markerHeight="9" refX="8" refY="3" orient="auto" markerUnits="strokeWidth"><path d="M0,0 L0,6 L9,3 Z" fill="#7c8a9a"/></marker>\n' \
          '</defs>\n' + '\n'.join(elements) + '\n</svg>'

# ========================================
# SVG #2: Support Systems (13 systems)
# ========================================
SUP_W = 2200
CARD_W = 680
CARD_HEAD_H = 65
SUP_STEP_H = 105
SUP_STEP_GAP = 10
SUP_PAD_L = 25

# Group support steps by SOP prefix
sup_groups = {}
for s in sup_steps:
    p = s["sop"][0]
    sup_groups.setdefault(p, []).append(s)

system_meta = {
    "J": ("J", "บำรุงรักษาเครื่องจักร + Calibration", "🔧", "#C0392B", "#FADBD8", "critical"),
    "K": ("K", "ความปลอดภัย / 5S / EHS", "🦺", "#C0392B", "#FADBD8", "critical"),
    "L": ("L", "ฝึกอบรม + Competency Matrix", "🎓", "#C0392B", "#FADBD8", "critical"),
    "N": ("N", "ควบคุมเอกสาร SOP", "📚", "#C0392B", "#FADBD8", "critical"),
    "O": ("O", "KPI Dashboard", "📊", "#D68910", "#FCF3CF", "medium"),
    "P": ("P", "Internal Audit + Management Review", "🔍", "#D68910", "#FCF3CF", "medium"),
    "Q": ("Q", "พัฒนา Supplier", "🤝", "#D68910", "#FCF3CF", "medium"),
    "R": ("R", "Forecasting / S&OP", "📈", "#D68910", "#FCF3CF", "medium"),
    "S": ("S", "Warranty Management", "🛡️", "#D68910", "#FCF3CF", "medium"),
    "T": ("T", "Continuous Improvement / Kaizen", "💡", "#27AE60", "#D5F5E3", "nice"),
    "U": ("U", "HR + Payroll Integration", "👥", "#27AE60", "#D5F5E3", "nice"),
    "V": ("V", "IT Backup + BCP", "💾", "#27AE60", "#D5F5E3", "nice"),
    "W": ("W", "Risk Register", "⚠️", "#27AE60", "#D5F5E3", "nice"),
}

def render_sys_card(x, y, code):
    code_, name, icon, color_main, color_light, _tier = system_meta[code]
    steps = sup_groups[code]
    n = len(steps)
    card_h = CARD_HEAD_H + n * (SUP_STEP_H + SUP_STEP_GAP) + 15
    out = []
    out.append(f'<rect x="{x}" y="{y}" width="{CARD_W}" height="{card_h}" fill="#FFFFFF" stroke="{color_main}" stroke-width="2.5" rx="8"/>')
    out.append(f'<rect x="{x}" y="{y}" width="{CARD_W}" height="{CARD_HEAD_H}" fill="{color_main}" rx="8"/>')
    out.append(f'<rect x="{x}" y="{y+CARD_HEAD_H-12}" width="{CARD_W}" height="12" fill="{color_main}"/>')
    out.append(text(x+22, y+30, icon, 22, "bold", "start", "#FFFFFF"))
    out.append(text(x+58, y+28, f"{code_}.  {name}", 16, "bold", "start", "#FFFFFF"))
    out.append(text(x+58, y+48, f"{n} ขั้นตอน · ทำเป็นรอบ (Cyclical)", 11, "normal", "start", "#FFFFFFCC"))
    out.append(text(x+CARD_W-15, y+38, "🔄 วงจร", 13, "bold", "end", "#FFFFFFCC"))
    cy = y + CARD_HEAD_H + 8
    for step in steps:
        sx = x + 14
        sw = CARD_W - 28
        out.extend(render_box(sx, cy, sw, SUP_STEP_H, step, title_size=12))
        cy += SUP_STEP_H + SUP_STEP_GAP
    return out, card_h

elements = []
# Top header
sup_top_h = 100
elements.append(f'<rect x="0" y="0" width="{SUP_W}" height="{sup_top_h}" fill="#1F4E78"/>')
elements.append(text(SUP_W/2, 32, "ระบบสนับสนุน 13 ระบบ (67 ขั้น)", 22, "bold", "middle", "#FFFFFF"))
elements.append(text(SUP_W/2, 58, "ทำเป็นรอบ คู่ขนานกับงานประจำวัน · กดที่กล่องใดๆ เพื่อดูรายละเอียดเต็ม", 13, "normal", "middle", "#D6E4F0"))
elements.append(text(SUP_W/2, 80, "🔴 4 ระบบ Critical · 🟡 5 ระบบกลาง · 🟢 4 ระบบเสริม", 12, "normal", "middle", "#D6E4F0"))

def add_tier(y, label, color):
    elements.append(f'<rect x="{SUP_PAD_L-5}" y="{y}" width="{SUP_W-2*SUP_PAD_L+10}" height="35" fill="{color}" rx="4"/>')
    elements.append(text(SUP_PAD_L+15, y+23, label, 16, "bold", "start", "#FFFFFF"))

tier_order = [
    ("critical", "🔴 Critical — ขาดสำคัญ ส่งผลทันที (ต้องทำก่อน)", "#C0392B"),
    ("medium", "🟡 Medium — ขาดระดับกลาง ทำให้ระบบสมบูรณ์", "#D68910"),
    ("nice", "🟢 Nice to Have — ขาดระดับเสริม (ทำเพิ่มเพื่อสมบูรณ์)", "#27AE60"),
]
codes_by_tier = {"critical":[], "medium":[], "nice":[]}
for code, meta in system_meta.items():
    codes_by_tier[meta[5]].append(code)

tier_y = sup_top_h + 20
for tier, label, color in tier_order:
    add_tier(tier_y, label, color)
    col_h = [tier_y + 50, tier_y + 50]
    for idx, code in enumerate(codes_by_tier[tier]):
        col = idx % 2
        x = SUP_PAD_L + col * 720
        y = col_h[col]
        card_els, card_h = render_sys_card(x, y, code)
        elements.extend(card_els)
        col_h[col] += card_h + 25
    tier_y = max(col_h) + 25

SUP_H = tier_y + 50
svg_sup = f'<svg viewBox="0 0 {SUP_W} {SUP_H}" xmlns="http://www.w3.org/2000/svg" preserveAspectRatio="xMidYMid meet">\n' \
          '<defs>\n' \
          '<marker id="arr" markerWidth="9" markerHeight="9" refX="8" refY="3" orient="auto" markerUnits="strokeWidth"><path d="M0,0 L0,6 L9,3 Z" fill="#2c3e50"/></marker>\n' \
          '</defs>\n' + '\n'.join(elements) + '\n</svg>'

# ========================================
# Build searchable list HTML
# ========================================
def make_search_table():
    rows = []
    for s in all_steps:
        kind_label = {"proc":"ปกติ","ctrl":"⚑ ตรวจสอบ","gate":"⛔ ด่าน","artifact":"📋 เอกสาร"}[s["kind"]]
        kind_class = s["kind"]
        rows.append(f'''<tr class="search-row" data-step="{s["num"]}" data-text="{esc(s["sop"]+" "+s["title"]+" "+s["role"])}">
  <td class="num">{s["num"]}</td>
  <td class="sop">{esc(s["sop"])}</td>
  <td><span class="kind-badge kind-{kind_class}">{kind_label}</span></td>
  <td class="title">{esc(s["title"])}</td>
  <td class="role">{esc(s["role"])}</td>
</tr>''')
    return '\n'.join(rows)

# ========================================
# Single HTML
# ========================================
steps_json = json.dumps({s["num"]: s for s in all_steps}, ensure_ascii=False)

html = f'''<!DOCTYPE html>
<html lang="th">
<head>
<meta charset="UTF-8">
<title>CHH SOP — ระบบครบทุกขั้น 157 ขั้น (Interactive)</title>
<style>
* {{ box-sizing: border-box; }}
body {{
  margin: 0; padding: 0;
  font-family: "Sarabun", "Tahoma", "Arial", sans-serif;
  color: #1a2332; background: #eef2f7;
  line-height: 1.5;
}}
header {{
  background: linear-gradient(135deg, #1F4E78, #2E75B6);
  color: #fff; padding: 22px 30px;
  position: sticky; top: 0; z-index: 100;
  box-shadow: 0 2px 10px rgba(0,0,0,.15);
}}
header h1 {{ margin: 0; font-size: 22px; }}
header p {{ margin: 4px 0 0; font-size: 13px; color: #D6E4F0; }}
nav.tabs {{
  background: #fff;
  padding: 0 30px;
  border-bottom: 1px solid #d8dde3;
  position: sticky; top: 80px; z-index: 99;
  box-shadow: 0 1px 3px rgba(0,0,0,.05);
  display: flex; gap: 4px;
}}
.tab-btn {{
  background: transparent; border: none; cursor: pointer;
  padding: 14px 22px; font-size: 14px; font-weight: 600;
  color: #5a6b7d; border-bottom: 3px solid transparent;
  font-family: inherit;
}}
.tab-btn:hover {{ color: #1F4E78; background: #f6f8fb; }}
.tab-btn.active {{
  color: #1F4E78; border-bottom-color: #2E75B6;
  background: #f6f8fb;
}}
main {{ padding: 18px 30px 40px; }}
.tab-content {{ display: none; }}
.tab-content.active {{ display: block; }}

.stats {{ display: flex; gap: 12px; margin-bottom: 16px; flex-wrap: wrap; }}
.stat {{
  background: #fff; padding: 12px 18px; border-radius: 8px;
  box-shadow: 0 1px 3px rgba(0,0,0,.06); flex: 1; min-width: 130px;
}}
.stat-num {{ font-size: 24px; font-weight: bold; color: #1F4E78; }}
.stat-label {{ font-size: 12px; color: #5a6b7d; }}

.legend {{
  background: #fff; border-radius: 8px; padding: 12px 18px;
  margin-bottom: 14px; box-shadow: 0 1px 3px rgba(0,0,0,.06);
  font-size: 13px;
  display: flex; flex-wrap: wrap; gap: 14px;
  align-items: center;
}}
.legend-item {{ display: flex; align-items: center; gap: 6px; }}
.swatch {{ width: 22px; height: 14px; border-radius: 3px; display: inline-block; }}

.diagram-wrap {{
  background: #fff; border-radius: 8px; padding: 14px;
  box-shadow: 0 1px 3px rgba(0,0,0,.06); overflow: auto;
  max-height: 80vh;
}}
svg {{ display: block; }}

.step-box:hover rect:first-of-type {{
  filter: brightness(1.05) drop-shadow(0 2px 6px rgba(0,0,0,0.18));
  transition: filter 0.15s;
}}

/* Modal */
.modal-backdrop {{
  position: fixed; inset: 0; background: rgba(0,0,0,0.55);
  display: none; align-items: flex-start; justify-content: center;
  z-index: 1000; padding: 40px 20px; overflow-y: auto;
  animation: fadeIn 0.15s;
}}
.modal-backdrop.show {{ display: flex; }}
.modal {{
  background: #fff; border-radius: 12px; max-width: 900px; width: 100%;
  box-shadow: 0 10px 40px rgba(0,0,0,0.3);
  animation: slideUp 0.2s;
  overflow: hidden;
  margin-bottom: 40px;
}}
@keyframes fadeIn {{ from {{ opacity: 0 }} to {{ opacity: 1 }} }}
@keyframes slideUp {{ from {{ opacity: 0; transform: translateY(20px) }} to {{ opacity: 1; transform: translateY(0) }} }}

.modal-header {{
  background: linear-gradient(135deg, #1F4E78, #2E75B6);
  color: #fff; padding: 22px 28px;
  display: flex; justify-content: space-between; align-items: center;
}}
.modal-header.kind-ctrl {{ background: linear-gradient(135deg, #E65100, #F57F17); }}
.modal-header.kind-gate {{ background: linear-gradient(135deg, #B71C1C, #C62828); }}
.modal-header.kind-artifact {{ background: linear-gradient(135deg, #4A148C, #6A1B9A); }}
.modal-header h2 {{ margin: 0; font-size: 20px; }}
.modal-header .meta {{ font-size: 13px; opacity: 0.9; margin-top: 4px; }}
.modal-close {{
  background: rgba(255,255,255,0.2); border: none; color: #fff;
  width: 36px; height: 36px; border-radius: 50%; cursor: pointer;
  font-size: 22px; font-weight: bold; transition: background 0.15s;
}}
.modal-close:hover {{ background: rgba(255,255,255,0.35); }}

.modal-body {{ padding: 24px 28px; }}
.modal-section-grid {{
  display: grid; gap: 14px;
  grid-template-columns: 1fr 1fr;
  margin-bottom: 16px;
}}
.modal-section {{
  background: #f6f8fb; border-left: 4px solid #2E75B6;
  padding: 12px 16px; border-radius: 6px;
}}
.modal-section.input {{ border-left-color: #1976D2; background: #E3F2FD; }}
.modal-section.data {{ border-left-color: #7B1FA2; background: #F3E5F5; }}
.modal-section.process {{ border-left-color: #F57F17; background: #FFF8E1; grid-column: span 2; }}
.modal-section.output {{ border-left-color: #2E7D32; background: #E8F5E9; grid-column: span 2; }}
.modal-section h3 {{
  margin: 0 0 6px; font-size: 13px; font-weight: 700;
  text-transform: uppercase; letter-spacing: 0.4px; color: #1a2332;
}}
.modal-section p, .modal-section .lines {{
  margin: 0; font-size: 14px; color: #1a2332; white-space: pre-wrap;
}}
.modal-section .lines li {{ margin: 4px 0; }}

.modal-meta-grid {{
  display: grid; grid-template-columns: 1fr 1fr; gap: 10px;
  background: #f6f8fb; padding: 14px 18px; border-radius: 6px;
  border-top: 3px solid #d8dde3;
}}
.modal-meta-grid > div {{ font-size: 13px; }}
.modal-meta-grid > div strong {{ color: #1F4E78; }}

.modal-note {{
  margin-top: 12px;
  background: #FFF3CD; border-left: 4px solid #A47D14;
  padding: 10px 14px; border-radius: 6px; font-size: 13px;
}}

/* Search tab */
.search-input {{
  width: 100%; padding: 12px 16px; font-size: 15px;
  border: 1.5px solid #d8dde3; border-radius: 8px;
  font-family: inherit; margin-bottom: 12px;
}}
.search-input:focus {{ outline: none; border-color: #2E75B6; }}
table.search-table {{
  width: 100%; background: #fff; border-radius: 8px;
  border-collapse: collapse; box-shadow: 0 1px 3px rgba(0,0,0,.06);
  overflow: hidden;
}}
.search-table th {{
  background: #1F4E78; color: #fff; padding: 12px;
  text-align: left; font-size: 13px;
}}
.search-table td {{
  padding: 10px 12px; border-bottom: 1px solid #eef2f7;
  font-size: 13px;
}}
.search-row {{ cursor: pointer; transition: background 0.1s; }}
.search-row:hover {{ background: #f0f7fc; }}
.search-row .num {{ font-weight: bold; color: #1F4E78; width: 60px; }}
.search-row .sop {{ font-family: monospace; color: #5d4037; width: 70px; }}
.kind-badge {{
  display: inline-block; padding: 2px 8px; border-radius: 12px;
  font-size: 11px; font-weight: bold;
}}
.kind-proc {{ background: #FFF3CD; color: #5D4037; }}
.kind-ctrl {{ background: #FFF9C4; color: #E65100; }}
.kind-gate {{ background: #FFEBEE; color: #B71C1C; }}
.kind-artifact {{ background: #E1BEE7; color: #4A148C; }}

@media print {{
  nav.tabs, .modal-backdrop, header {{ display: none; }}
  .tab-content {{ display: block !important; page-break-after: always; }}
  body {{ background: #fff; }}
}}
</style>
</head>
<body>

<header>
  <h1>CHH SOP — ระบบทำงานครบทั้งหมด 157 ขั้น</h1>
  <p>งานประจำวัน 90 ขั้น (Phase 1-6) + ระบบสนับสนุน 67 ขั้น (13 ระบบ) — กดที่กล่องใดก็ได้ดูรายละเอียดเต็ม</p>
</header>

<nav class="tabs">
  <button class="tab-btn active" data-tab="ops">📋 งานประจำวัน 90 ขั้น</button>
  <button class="tab-btn" data-tab="sup">🔄 ระบบสนับสนุน 13 ระบบ</button>
  <button class="tab-btn" data-tab="search">🔍 ค้นหา 157 ขั้น</button>
</nav>

<main>

<section id="tab-ops" class="tab-content active">
  <div class="stats">
    <div class="stat"><div class="stat-num">90</div><div class="stat-label">ขั้นตอน 1-90</div></div>
    <div class="stat"><div class="stat-num">6</div><div class="stat-label">Phase หลัก</div></div>
    <div class="stat"><div class="stat-num">12</div><div class="stat-label">⚑ จุดตรวจสอบ</div></div>
    <div class="stat"><div class="stat-num">2</div><div class="stat-label">⛔ ด่านห้ามข้าม</div></div>
    <div class="stat"><div class="stat-num">1</div><div class="stat-label">📋 เอกสารหลัก</div></div>
  </div>
  <div class="legend">
    <div class="legend-item"><span class="swatch" style="background:#FFF3CD; border:1.5px solid #A47D14"></span>ขั้นตอนปกติ</div>
    <div class="legend-item"><span class="swatch" style="background:#FFF9C4; border:3px solid #F57F17"></span>⚑ จุดตรวจสอบ</div>
    <div class="legend-item"><span class="swatch" style="background:#FFEBEE; border:3px solid #C62828"></span>⛔ ด่านห้ามข้าม</div>
    <div class="legend-item"><span class="swatch" style="background:#E1BEE7; border:1.5px solid #6A1B9A; border-radius:8px"></span>📋 เอกสารหลัก</div>
    <div class="legend-item" style="margin-left:auto; color:#1F4E78; font-weight:bold;">💡 กดที่กล่องเพื่อดูรายละเอียด</div>
  </div>
  <div class="diagram-wrap">
    {svg_ops}
  </div>
</section>

<section id="tab-sup" class="tab-content">
  <div class="stats">
    <div class="stat"><div class="stat-num" style="color:#C0392B">4</div><div class="stat-label">🔴 ระบบ Critical</div></div>
    <div class="stat"><div class="stat-num" style="color:#D68910">5</div><div class="stat-label">🟡 ระบบกลาง</div></div>
    <div class="stat"><div class="stat-num" style="color:#27AE60">4</div><div class="stat-label">🟢 ระบบเสริม</div></div>
    <div class="stat"><div class="stat-num">13</div><div class="stat-label">รวมระบบ</div></div>
    <div class="stat"><div class="stat-num">67</div><div class="stat-label">ขั้นที่ 91-157</div></div>
  </div>
  <div class="legend">
    <div class="legend-item">ระบบสนับสนุน = ทำเป็นรอบ (Cyclical) คู่ขนานกับงานประจำวัน</div>
    <div class="legend-item" style="margin-left:auto; color:#1F4E78; font-weight:bold;">💡 กดที่กล่องเพื่อดูรายละเอียด</div>
  </div>
  <div class="diagram-wrap">
    {svg_sup}
  </div>
</section>

<section id="tab-search" class="tab-content">
  <div class="stats">
    <div class="stat"><div class="stat-num">157</div><div class="stat-label">ขั้นตอนทั้งหมด</div></div>
    <div class="stat"><div class="stat-num">19</div><div class="stat-label">หมวด/ระบบ</div></div>
  </div>
  <input type="text" class="search-input" id="searchInput" placeholder="🔍 พิมพ์เพื่อค้นหา (รหัส SOP, ชื่อขั้นตอน, ผู้รับผิดชอบ, ฯลฯ)">
  <table class="search-table">
    <thead>
      <tr><th>ขั้นที่</th><th>SOP</th><th>ประเภท</th><th>ชื่อขั้นตอน</th><th>ผู้รับผิดชอบ</th></tr>
    </thead>
    <tbody>
      {make_search_table()}
    </tbody>
  </table>
</section>

</main>

<!-- Modal -->
<div class="modal-backdrop" id="modal">
  <div class="modal" id="modalContent"></div>
</div>

<script>
const STEPS = {steps_json};

// Tab switching
document.querySelectorAll('.tab-btn').forEach(btn => {{
  btn.addEventListener('click', () => {{
    document.querySelectorAll('.tab-btn').forEach(b => b.classList.remove('active'));
    document.querySelectorAll('.tab-content').forEach(c => c.classList.remove('active'));
    btn.classList.add('active');
    document.getElementById('tab-' + btn.dataset.tab).classList.add('active');
    window.scrollTo({{top:0, behavior:'smooth'}});
  }});
}});

// Modal logic
function fmtProcess(text) {{
  if (!text || text === '—') return '<em style="color:#999">ไม่มีข้อมูล</em>';
  // detect numbered steps "1) ... 2) ..."
  const lines = text.split(/\\n|(?=\\d+\\))/).filter(s => s.trim());
  if (lines.length > 1 && lines[0].match(/^\\d+\\)/)) {{
    return '<ol class="lines" style="margin:0; padding-left:18px;">' +
      lines.map(l => '<li>' + escapeHtml(l.replace(/^\\d+\\)\\s*/, '')) + '</li>').join('') +
      '</ol>';
  }}
  return '<p>' + escapeHtml(text).replace(/\\n/g, '<br>') + '</p>';
}}
function escapeHtml(s) {{
  return String(s).replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;');
}}
function showStep(num) {{
  const s = STEPS[num];
  if (!s) return;
  const kindLabel = {{proc:"ขั้นตอนปกติ", ctrl:"⚑ จุดตรวจสอบสำคัญ", gate:"⛔ ด่านห้ามข้าม", artifact:"📋 เอกสารหลัก"}}[s.kind];
  document.getElementById('modalContent').innerHTML = `
    <div class="modal-header kind-${{s.kind}}">
      <div>
        <h2>${{escapeHtml(s.title)}}</h2>
        <div class="meta">ขั้นที่ ${{s.num}} · รหัส ${{escapeHtml(s.sop)}} · ${{kindLabel}}</div>
      </div>
      <button class="modal-close" onclick="closeModal()">×</button>
    </div>
    <div class="modal-body">
      <div class="modal-section-grid">
        <div class="modal-section input">
          <h3>📥 Input — สิ่งที่รับเข้ามา</h3>
          <p>${{escapeHtml(s.input)}}</p>
        </div>
        <div class="modal-section data">
          <h3>📊 Data — ข้อมูลที่จัดการ</h3>
          <p>${{escapeHtml(s.data)}}</p>
        </div>
        <div class="modal-section process">
          <h3>⚙️ Process — ขั้นตอนการทำงาน</h3>
          ${{fmtProcess(s.process)}}
        </div>
        <div class="modal-section output">
          <h3>📤 Output — สิ่งที่ส่งออก</h3>
          <p>${{escapeHtml(s.output)}}</p>
        </div>
      </div>
      <div class="modal-meta-grid">
        <div><strong>👤 ผู้รับผิดชอบ:</strong><br>${{escapeHtml(s.role)}}</div>
        <div><strong>✍ ผู้อนุมัติ:</strong><br>${{escapeHtml(s.approver)}}</div>
        <div style="grid-column: span 2"><strong>🛠 เครื่องมือที่ใช้:</strong> ${{escapeHtml(s.tool)}}</div>
        ${{s.section ? `<div style="grid-column: span 2; color:#5a6b7d; font-size:12px;"><strong>📂 หมวด:</strong> ${{escapeHtml(s.section)}}</div>` : ''}}
      </div>
      ${{s.note && s.note !== '—' ? `<div class="modal-note">💡 <strong>หมายเหตุสำคัญ:</strong> ${{escapeHtml(s.note)}}</div>` : ''}}
    </div>
  `;
  document.getElementById('modal').classList.add('show');
}}
function closeModal() {{
  document.getElementById('modal').classList.remove('show');
}}

// Click handlers (delegated)
document.addEventListener('click', (e) => {{
  const box = e.target.closest('.step-box, .search-row');
  if (box) {{
    showStep(box.dataset.step);
  }}
}});
document.getElementById('modal').addEventListener('click', (e) => {{
  if (e.target.id === 'modal') closeModal();
}});
document.addEventListener('keydown', (e) => {{
  if (e.key === 'Escape') closeModal();
}});

// Search
const searchInput = document.getElementById('searchInput');
if (searchInput) {{
  searchInput.addEventListener('input', () => {{
    const q = searchInput.value.toLowerCase().trim();
    document.querySelectorAll('.search-row').forEach(row => {{
      const txt = row.dataset.text.toLowerCase();
      row.style.display = !q || txt.includes(q) ? '' : 'none';
    }});
  }});
}}
</script>

</body>
</html>'''

with open("/Users/panuwatjangchudjai/sop/CHH_SOP_Complete.html", "w", encoding="utf-8") as f:
    f.write(html)

print(f"OK — generated CHH_SOP_Complete.html")
print(f"  - Operations: {len(ops_steps)} steps")
print(f"  - Support: {len(sup_steps)} steps")
print(f"  - Total: {len(all_steps)} steps")
